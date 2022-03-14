from openpyxl import load_workbook

def avg_1_round(wb, team_name, sheets):
    total_points = 0
    appereance = 0
    for sheet in sheets:
        ws = wb[sheet]

        for column in range(2, 12):
            if team_name == ws.cell(row=1, column=column).value:
                point = ws.cell(row=10, column=column).value
                total_points += point
                appereance += 1
                break
    avg = total_points/appereance
    print(f'La media dopo 1 girone è {avg}')
    return avg, team_name

def avg_2_round(wb, team_name, sheets):
    total_points = 0
    appereance = 0
    for sheet in sheets:
        ws = wb[sheet]

        for column in range(2, 12):
            if team_name == ws.cell(row=1, column=column).value:
                point = ws.cell(row=19, column=column).value
                total_points += point
                appereance += 1
                break

    avg = total_points / appereance
    print(f'La media dopo 2 gironi è {avg}')
    return avg, team_name

def avg_3_round(wb, team_name, sheets):
    total_points = 0
    appereance = 0
    for sheet in sheets:
        ws = wb[sheet]

        for column in range(2, 12):
            if team_name == ws.cell(row=1, column=column).value:
                point = ws.cell(row=28, column=column).value
                total_points += point
                appereance += 1
                break

    avg = total_points / appereance
    print(f'La media dopo 3 gironi è {avg}')
    return avg, team_name

def avg_4_round(wb, team_name, sheets):
    total_points = 0
    appereance = 0
    for sheet in sheets:
        ws = wb[sheet]

        for column in range(2, 12):
            if team_name == ws.cell(row=1, column=column).value:
                point = ws.cell(row=37, column=column).value
                total_points += point
                appereance += 1
                break

    avg = total_points / appereance
    print(f'La media dopo 4 gironi è {avg}')
    print(f'Il numero totale di punti è: {total_points}')
    print(f'Il numero totale di apparizioni è: {appereance}')
    return avg, appereance

def fill_info(wb, team_name, avg_1, avg_2, avg_3, avg_4, app):
    ws = wb['Statistiche dal 2014-2015']
    for row in range(8, 27):
        for column in range(1, 17):
            if team_name == ws.cell(row, column).value:
                ws.cell(row+1, column+1).value = avg_1
                ws.cell(row+2, column + 1).value = avg_2
                ws.cell(row+3, column + 1).value = avg_3
                ws.cell(row+4, column + 1).value = avg_4
                ws.cell(row+6, column + 1).value = app



wb = load_workbook('/Users/marcomelis/Dropbox/Mia/File excel/Fantacalcio/Storico/Storico anni precedenti.xlsx')
sheets = wb.sheetnames[3:-1]
teams_name = ['Bisca', 'Carlo', 'Giulio', 'Melis', 'Michelone', 'Milo', 'Nathan', 'Pietro', 'Sangio', 'Zino', 'Floro', 'Gracis', 'Zone']
for team_name in teams_name:
    print(team_name)
    avg_1,_ = avg_1_round(wb, team_name, sheets)
    avg_2,_ = avg_2_round(wb, team_name, sheets)
    avg_3,_ = avg_3_round(wb, team_name, sheets)
    avg_4, app = avg_4_round(wb, team_name, sheets)
    fill_info(wb, team_name, avg_1, avg_2, avg_3, avg_4, app)

wb.save('/Users/marcomelis/Dropbox/Mia/File excel/Fantacalcio/Storico/Storico anni precedenti.xlsx')
wb.close()
