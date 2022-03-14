from openpyxl import load_workbook

# wb = load_workbook('/Users/marcomelis/Dropbox/Mia/File excel/Fantacalcio/2019-2020/Fantacalcio 2019-2020.xlsx', data_only=True)
# ws = wb['Calendario']
# counter = 0
# for row in range (2, 182):
#     number_match = (row -2)//5 + 1
#     team_home = ws.cell(row, 2).value
#     points_home = ws.cell(row, 4).value
#     modifier_home = ws.cell(row, 5).value
#     team_away = ws.cell(row, 9).value
#     points_away = ws.cell(row, 7).value
#     modifier_away = ws.cell(row, 6).value
#
#     if points_home is not None and float(points_home) // 6 != float(points_home - modifier_away) // 6:
#         print(f'Giornata {number_match}: {team_home}-{team_away} cambia risultato')
#         if (points_home - 60) // 6
#         counter += 1
#     if points_away is not None and float(points_away) // 6 != float(points_away - modifier_home) // 6:
#         print(f'Giornata {number_match}: {team_home}-{team_away} cambia risultato')
#         counter += 1
# print(f'Risultato cambiato {counter} volte')

wb = load_workbook('/Users/marcomelis/Dropbox/Mia/File excel/Fantacalcio/2020-2021/Fantacalcio 2020-2021.xlsx')
ws_calendario = wb['Calendario senza modificatore']
for row in range(2, 132):
    home_points = ws_calendario.cell(row, 4).value
    away_points = ws_calendario.cell(row, 7).value
    ws_calendario.cell(row, 4).value = f'{home_points[:-1]} - F{row})'
    ws_calendario.cell(row, 7).value = f'{away_points[:-1]} - E{row})'
    print('OK')

wb.save('/Users/marcomelis/Dropbox/Mia/File excel/Fantacalcio/2020-2021/Fantacalcio 2020-2021.xlsx')
wb.close()