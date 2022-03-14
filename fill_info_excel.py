from json import load

from openpyxl import load_workbook

from COSTANTS import CURRENT_YEAR, FILE_PATH
from add_link import *


def fill_info_excel(wb, match_day):

    # open json detail file
    input_file = open(f'resource/{CURRENT_YEAR}/Formazioni/Giornata' + match_day + '.json', 'r')
    json_input_file = load(input_file)
    team_name = list(json_input_file['Giornata' + match_day])

    # Defines the minimum row and columns for each match
    row = [4, 24, 44, 64, 84]
    column = [1, 8]

    # Load thw workbook
    #wb = load_workbook(f'{FILE_PATH}')

    # Check if the worksheet is present into the workbook
    if "Giornata " + match_day in wb.sheetnames:
        print("Il foglio è presente")
        # Assign the right worksheet
        ws = wb["Giornata " + match_day]
    else:
        print("Il foglio non è presente. Creazione nuovo foglio...")
        foglio_guida = wb['Foglio guida']
        ws = wb.copy_worksheet(foglio_guida)
        ws.title = 'Giornata ' + match_day

    prepare_spreadsheet(wb, match_day)

    num_player = 0
    counter_team_name = 0
    for k in range(5):
        row_line = row[k]
        for index in range(1, 12):
            ws.cell(row_line, column[0]).value = json_input_file['Giornata' + match_day][str(team_name[counter_team_name])]['titolari'][str(index)]
            ws.cell(row_line, column[1]).value = json_input_file['Giornata' + match_day][str(team_name[counter_team_name + 1])]['titolari'][str(index)]
            row_line += 1
            num_player += 1
        counter_team_name += 2

    # add_link(ws, match_links)
    add_link(ws, json_input_file)

    ws = wb["Sostituti"]
    row = [4, 22, 40, 58, 76]
    column = [1, 5]
    counter_team_name = 0
    for times in range(5):
        row_line = row[times]
        for index in range(1, 15):
            ws.cell(row_line, column[0]).value = json_input_file['Giornata' + match_day][str(team_name[counter_team_name])]['riserve'][str(index)]
            ws.cell(row_line, column[1]).value = json_input_file['Giornata' + match_day][str(team_name[counter_team_name + 1])]['riserve'][str(index)]
            row_line += 1
        counter_team_name += 2

    ws = wb['Formazioni non schierate']
    for times in range(2, 12):
        player = ws.cell(times, 1).value
        ws.cell(times, int(match_day) + 1).value = json_input_file['Giornata' + match_day][player]['formazione schierata']
        if json_input_file['Giornata' + match_day][player]['formazione schierata'] == 'X':
            print(player + ' non ha schierato la formazione')

    wb.save(f'{FILE_PATH}')
    wb.close()

def prepare_spreadsheet(wb, match_day):
    set_team_name(wb, match_day)
    delete_substitute_player(wb)
    replace_giornata_string_sostituti(wb, match_day)
    replace_teams_name_substitute_worksheet(wb, match_day)
    set_calendar(wb, match_day)
    set_my_calendar(wb, match_day)

def delete_substitute_player(wb):
    ws = wb['Sostituti']
    columns = [1, 5]
    for column in columns:
        for row in range(1, 90):
            cell_value = ws.cell(row, column).value
            if cell_value is not None and cell_value != 'Giocatore' and 'Partita' not in cell_value:
                ws.cell(row, column).value = ''

    print('Cancellazione sostituti partita precedente...')

def replace_giornata_string_sostituti(wb, match_day):
    # Assign the worksheet
    ws_sostituti = wb['Sostituti']
    # set variable with the number of last played matches
    last_giornata = str(int(match_day) - 1)
    columns = [3, 7]
    # loop through column list
    for column in columns:
        # loop through row in specified range
        for row in range(4, 90):
            # check if value cell is not None
            if ws_sostituti.cell(row, column).value is not None:
                # take the value of specific cell and replace old data with new data (new match_day)
                new_value = ws_sostituti.cell(row, column).value.replace('Giornata ' + last_giornata, 'Giornata ' + match_day)
                # overwrite value cell
                ws_sostituti.cell(row, column).value = new_value

    print('Cambio formule foglio \'Sostituti\'')

def set_calendar(wb, match_day):

    ws = wb['Calendario']
    for row in range(2, 178, 5):
        if ws.cell(row, 1).value == 'Giornata ' + match_day:
            start_row = row
            break
    paste_formulae_calendar(start_row, start_row + 5, ws, match_day)

    print('Copia e cambio formule foglio \'Calendario\'')

def paste_formulae_calendar(start_row, end_row, ws, match_day):

    formulae = [
        '=IF(ISERROR(\'Giornata MATCHDAY\'!$P$33), "", VLOOKUP(CELL,\'Giornata MATCHDAY\'!$O$25:$R$34,2,FALSE))',
        '=IF(ISERROR(\'Giornata MATCHDAY\'!$Q$33), "", VLOOKUP(CELL,\'Giornata MATCHDAY\'!$O$25:$R$34,3,FALSE))',
        '=VLOOKUP(CELL,\'Giornata MATCHDAY\'!$O$25:$R$34,4,FALSE)'

        ]
    counter = 0

    # copy in copied_cell list the value of each cell between start_row-end_row and start_column-end_column
    for row in range(start_row, end_row):
        for column in range(3, 6):
            ws.cell(row, column).value = formulae[counter].replace('MATCHDAY', match_day).replace('CELL', 'B' + str(row))
            counter += 1
        counter = 2
        for column in range(6, 9):
            ws.cell(row, column).value = formulae[counter].replace('MATCHDAY', match_day).replace('CELL', 'I' + str(row))
            counter -= 1
        counter = 0

def set_my_calendar(wb, match_day):

    ws = wb['Mio calendario']
    formulae = [
        '=VLOOKUP(CELL,\'Giornata MATCHDAY\'!$O$25:$R$34,2,FALSE)',
        '=VLOOKUP(CELL,\'Giornata MATCHDAY\'!$O$25:$R$34,3,FALSE)',
    ]
    counter = 0

    for row in range(2, 38):
        if ws.cell(row, 7).value == f'Giornata {match_day}':
            main_row = row
            break
    for column in range(2, 4):
        ws.cell(main_row, column).value = formulae[counter].replace('MATCHDAY', match_day).replace('CELL', 'A' + str(main_row))
        counter += 1
    counter = 1
    for column in range(4, 6):
        ws.cell(main_row, column).value = formulae[counter].replace('MATCHDAY', match_day).replace('CELL', 'F' + str(main_row))
        counter -= 1

    print('Cambio formule foglio \'Mio calendario\'')

def replace_teams_name_substitute_worksheet(wb, match_day):

    ws_sostituti = wb['Sostituti']
    ws_calendario = wb['Calendario']
    start_row = 2 + (int(match_day) - 1)*5

    for row in range(2, 75, 18):
        ws_sostituti.cell(row, 1).value = ws_calendario.cell(start_row, 2).value
        ws_sostituti.cell(row, 5).value = ws_calendario.cell(start_row, 9).value
        start_row += 1

def set_team_name(wb, match_day):

    ws_calendario = wb['Calendario']
    ws_day = wb[f'Giornata {match_day}']

    for row_calendario in range(2, 182, 5):
        current_day = ws_calendario.cell(row_calendario, 1).value
        if current_day == f'Giornata {match_day}':
            break

    for row in range(25, 35, 2):
        ws_day.cell(row, 15).value = ws_calendario.cell(row_calendario, 2).value
        row+=1
        ws_day.cell(row, 15).value = ws_calendario.cell(row_calendario, 9).value
        row_calendario+=1










