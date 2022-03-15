import openpyxl
import json

from sys import path
path.insert(0, r"C:\Users\mmelis\Downloads\Fantacalcio\Fantacalcio")

from src.COSTANTS import CURRENT_YEAR


def create_json_change_name():
    data = dict()
    wb = openpyxl.load_workbook(r"C:\Users\mmelis\Downloads\2021-2022\2021-2022\Associazione giocatori FantaService-PianetaFanta.xlsx")
    ws = wb['Foglio1']
    row_index = 2


    while ws.cell(row_index, 1).value is not None:
        print(f'Row: {row_index}')
        data[ws.cell(row_index, 1).value] = dict()
        data[ws.cell(row_index, 1).value]['nome_PianetaFanta'] = ws.cell(row_index, 3).value
        data[ws.cell(row_index, 1).value]['Ruolo'] = ws.cell(row_index, 5).value
        data[ws.cell(row_index, 1).value]['Squadra'] = ws.cell(row_index, 4).value
        row_index += 1

    with open(f'rsc/change_name/change_name_febbraio_{CURRENT_YEAR}.json', 'w', encoding='utf-8') as output_file:
        json.dump(data, output_file, indent=4, ensure_ascii=False)


create_json_change_name()
