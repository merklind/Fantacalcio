#! /usr/local/bin/python3.7

import openpyxl
from sys import exit
from src.COSTANTS import FILE_PATH

def doing_substitution():
    wb = openpyxl.load_workbook(f'{FILE_PATH}', data_only = True)
    wb_formula = openpyxl.load_workbook(f'{FILE_PATH}')
    match_line = ((4, 6), (4, 13), (24, 6), (24, 13), (44, 6), (44, 13), (64, 6), (64, 13), (84, 6), (84, 13))
    sub_line = ((4, 3), (4, 7), (22, 3), (22, 7), (40, 3), (40, 7), (58, 3), (58, 7), (76, 3), (76, 7))
    # Insert the match number
    match_day = input('Inserisci la giornata: ')
    # Check if the worksheet is present into the workbook
    if "Giornata " + match_day in wb.sheetnames:
        print("Il foglio è presente")
        # Assign the right worksheet
        ws_f_giornata = wb_formula["Giornata " + match_day]
        ws_giornata = wb["Giornata " + match_day]
        ws_substitute = wb["Sostituti"]
    else:
        print("Il foglio non è presente")
        wb.close()
        wb_formula.close()
        exit()

    for num_match in range(10):
        # number of substitution already done
        num_sub = 0
        # start row of the match
        row_match = match_line[num_match][0]
        # start column of the match
        column_match = match_line[num_match][1]
        # start row of substitution sheet for num_match match
        row_sub = sub_line[num_match][0]
        # start column of substitution sheet for num_match match
        column_sub = sub_line[num_match][1]
        # loop through
        for line_regular_player in range(11):
            if(ws_giornata.cell(row_match + line_regular_player, column_match).value == "#N/A" or ws_giornata.cell(row_match + line_regular_player, column_match).value == "#VALUE!"):
                role = ws_giornata.cell(row_match + line_regular_player, column_match - 4).value
                for line_bleacher_player in range(14):
                    if(ws_substitute.cell(row_sub + line_bleacher_player, column_sub).value == 'Sì' and ws_substitute.cell(row_sub + line_bleacher_player, column_sub - 1).value == role):
                        ws_f_giornata.cell(row_match + line_regular_player, column_match - 5).value = ws_substitute.cell(row_sub + line_bleacher_player, column_sub - 2).value + "*"
                        ws_substitute.cell(row_sub + line_bleacher_player, column_sub).value = "Sostituito"
                        num_sub += 1
                        break
                    if(num_sub == 3 or line_bleacher_player == 13):
                        ws_f_giornata.cell(row_match + line_regular_player, column_match - 4).value = role
                        ws_f_giornata.cell(row_match + line_regular_player, column_match - 5).value = "Non gioca"
                        break

    wb_formula.save(f'{FILE_PATH}')
    wb_formula.close()


doing_substitution()