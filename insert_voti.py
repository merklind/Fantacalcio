import openpyxl
import json
from COSTANTS import FILE_PATH, CURRENT_YEAR

def insert_voti(match_day):
    wb = openpyxl.load_workbook(f'{FILE_PATH}')
    input_file = open(f'/Users/marcomelis/PycharmProjects/Fantacalcio/resource/{CURRENT_YEAR}/Voti ufficiosi/Json/Giornata ' + str(match_day + 2) + '.json')
    data = json.load(input_file)

    ws = wb['Giornata ' + str(match_day)]
    start_row = 2
    start_column = 24
    key_list = list(data.keys())
    counter = 0

    for row in range(start_row, start_row + len(data)):
        ws.cell(row, start_column).value = key_list[counter]
        ws.cell(row, start_column + 1).value = data[key_list[counter]]['Ruolo']
        ws.cell(row, start_column + 2).value = data[key_list[counter]]['Squadra']
        ws.cell(row, start_column + 3).value = data[key_list[counter]]['G']
        ws.cell(row, start_column + 3).number_format = 'General'
        ws.cell(row, start_column + 4).value = data[key_list[counter]]['GF']
        ws.cell(row, start_column + 5).value = data[key_list[counter]]['GS']
        ws.cell(row, start_column + 6).value = data[key_list[counter]]['Aut']
        ws.cell(row, start_column + 7).value = data[key_list[counter]]['Ass']
        ws.cell(row, start_column + 8).value = data[key_list[counter]]['Amm']
        ws.cell(row, start_column + 9).value = data[key_list[counter]]['Esp']
        ws.cell(row, start_column + 10).value = data[key_list[counter]]['RigS']
        ws.cell(row, start_column + 11).value = data[key_list[counter]]['RigP']
        counter += 1

    wb.save(f'{FILE_PATH}')
    wb.close()
